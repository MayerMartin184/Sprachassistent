from types import SimpleNamespace

from sprachassistent.agent.agent import Agent, request_extras, server_tools
from sprachassistent.config import Settings
from sprachassistent.tools import documents
from sprachassistent.tools.base import ToolRegistry
from sprachassistent.tools.files import FileManager


def test_create_office_documents(tmp_path):
    fm = FileManager({"Dokumente": tmp_path})
    msg = documents.create_docx(fm, "Dokumente/Jarvis/Test", "Angebot", "# Umfang\n- Punkt **eins**\n- Punkt zwei\n\nAbsatz.\n1. Schritt")
    assert "Test.docx" in msg
    import docx
    d = docx.Document(str(tmp_path / "Jarvis/Test.docx"))
    texts = [p.text for p in d.paragraphs]
    assert "Angebot" in texts and "Punkt eins" in texts and "Schritt" in texts

    msg = documents.create_xlsx(fm, "Dokumente/Jarvis/Tabelle", [{"name": "Kosten", "rows": [["Posten", "Betrag"], ["Kabel", "12,5"], ["Dosen", "3"]]}])
    import openpyxl
    ws = openpyxl.load_workbook(str(tmp_path / "Jarvis/Tabelle.xlsx"))["Kosten"]
    assert ws["A1"].value == "Posten" and ws["B2"].value == 12.5 and ws["B3"].value == 3

    msg = documents.create_pptx(fm, "Dokumente/Jarvis/Deck", "Projekt X", "Kickoff", [{"title": "Ziele", "bullets": ["A", "B"], "notes": "n"}])
    from pptx import Presentation
    prs = Presentation(str(tmp_path / "Jarvis/Deck.pptx"))
    assert len(prs.slides) == 2 and "2 Folien" in msg

    try:
        documents.create_docx(fm, "Dokumente/Jarvis/Test", "x", "y")
        raise AssertionError("sollte nicht überschreiben")
    except FileExistsError:
        pass


def test_model_dependent_request_parameters():
    assert server_tools("claude-opus-5")[0]["type"] == "web_search_20260209"
    assert server_tools("claude-haiku-4-5")[0]["type"] == "web_search_20250305"
    assert request_extras("claude-opus-5", "high") == {"output_config": {"effort": "high"}}
    assert request_extras("claude-opus-5", "kaputt") == {"output_config": {"effort": "medium"}}
    assert request_extras("claude-haiku-4-5", "high") == {}


def test_agent_uses_override_model_and_ask():
    calls = []

    class Msgs:
        def create(self, **kw):
            calls.append(kw)
            return SimpleNamespace(stop_reason="end_turn", content=[SimpleNamespace(type="text", text="ok")])

    settings = Settings(_env_file=None, assistant_model="claude-opus-5", assistant_effort="medium")
    agent = Agent(settings, ToolRegistry(), client=SimpleNamespace(messages=Msgs()), model="claude-haiku-4-5")
    agent.run("hi")
    assert calls[-1]["model"] == "claude-haiku-4-5" and "output_config" not in calls[-1]
    assert agent.ask("claude-opus-5", "Frage", "xhigh") == "ok"
    assert calls[-1]["model"] == "claude-opus-5" and calls[-1]["output_config"] == {"effort": "xhigh"} and "tools" not in calls[-1]
